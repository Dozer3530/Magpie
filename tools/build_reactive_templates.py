"""Build the Reactive <Crop> templates from the Static ones (one-time / on change).

The Reactive feed (client-scattered points) is Survey123-only — no lab data —
and its locations come from the CSV, not fixed stakes. So each Reactive template
is its Static sibling with:

  * every LAB-page column removed (nutrients / rates / ratios / lab text), and
  * a `Field` column inserted right after `ID` (so the export can stamp
    ID = F1/F2..., Field = "Field F", Location = "lat, lon" from the CSV).

Data rows are cleared — the export rewrites rows 2..N each time. Styling is
preserved by copying the loaded workbook and only deleting/inserting columns.

Run from the repo root:  python tools/build_reactive_templates.py
Then eyeball and commit the two new xlsx.
"""
from __future__ import annotations

import sys
from copy import copy
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))  # repo root on path

import openpyxl

from app.config import PROJECT_ROOT
from app.crops import CROPS
from app.schema import page_of, read_template_fields


def _reactive_path(display_name: str) -> Path:
    return PROJECT_ROOT / f"Reactive {display_name} Template.xlsx"


def build_for(template_path: Path, out_path: Path) -> tuple[int, int]:
    fields = read_template_fields(template_path)
    lab_cols = sorted(
        (f.excel_col for f in fields if page_of(f) == "lab"), reverse=True
    )

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Drop lab columns (right-to-left so earlier indices stay valid).
    for col in lab_cols:
        ws.delete_cols(col, 1)

    # Insert the Field column after ID (col A); copy the ID header's styling.
    ws.insert_cols(2, 1)
    src, dst = ws.cell(row=1, column=1), ws.cell(row=1, column=2)
    dst.value = "Field"
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)

    # Clear pre-filled data rows — reactive points are written fresh on export.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    wb.save(out_path)
    return len(fields), len(lab_cols)


def main() -> None:
    print("Building Reactive templates from the Static ones\n")
    for crop in CROPS:
        out = _reactive_path(crop.display_name)
        total, dropped = build_for(crop.template_path, out)
        kept = total - dropped + 1  # +1 for the inserted Field column
        print(f"  {out.name}: {total} cols - {dropped} lab + Field = {kept} cols")
    print("\nDone. Eyeball the two files, then commit them.")


if __name__ == "__main__":
    main()
