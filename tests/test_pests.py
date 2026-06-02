"""Tests for the Pest ID feed: parser, service, export block, rename migration."""
from __future__ import annotations

import openpyxl
import pytest

from app.importers.pest import parse_pest_file
from app.schema import page_of, read_template_fields
from app.services import pests
from app.services import weeks as weeks_service

# Corn (L points): header has bug columns; week 1 has counts, week 2 is blank.
CORN_CSV = (
    "person,DATE,ID Number,CARD COMPLETED,Aphid,Lygus,fly spp.\n"
    ",June 1,L1_1,TRUE,,2,101\n"
    ",June 1,L2_1,TRUE,3,,113\n"
    ",June 1,L3_1,TRUE,,1,\n"
    ",June 8,L1_2,FALSE,,,\n"
    ",June 8,L2_2,FALSE,,,\n"
    ",,,\n"
)

# Canola (M points): living doc with no bug columns yet, all cards FALSE.
CANOLA_CSV = (
    "person,DATE,ID Number,CARD COMPLETED\n"
    ",June 1,M1_1,FALSE\n"
    ",June 1,M2_1,FALSE\n"
    ",June 8,M1_2,FALSE\n"
)


def _csv(tmp_path, name, text):
    p = tmp_path / name
    p.write_text(text, encoding="utf-8")
    return p


# ---- parser ----------------------------------------------------------------

def test_parse_detects_crop_and_weeks(tmp_path):
    parsed = parse_pest_file(_csv(tmp_path, "corn.csv", CORN_CSV))
    assert parsed.crop_code == "corn"
    assert parsed.bug_names == ["Aphid", "Lygus", "fly spp."]
    assert [w.index for w in parsed.weeks] == [1, 2]

    w1 = parsed.week(1)
    assert w1.date == "June 1"
    assert w1.cards_completed == 3
    assert set(w1.points) == {"L1", "L2", "L3"}
    assert w1.points["L1"].bugs == {"Lygus": "2", "fly spp.": "101"}
    assert set(w1.bug_types_present) == {"Aphid", "Lygus", "fly spp."}

    w2 = parsed.week(2)
    assert w2.cards_completed == 0
    assert w2.bug_types_present == []


def test_parse_canola_no_bugs(tmp_path):
    parsed = parse_pest_file(_csv(tmp_path, "canola.csv", CANOLA_CSV))
    assert parsed.crop_code == "canola"
    assert parsed.bug_names == []
    assert parsed.week(1).cards_completed == 0


def test_parse_rejects_mixed_prefix(tmp_path):
    bad = "p,DATE,ID Number,CARD COMPLETED\n,June 1,M1_1,TRUE\n,June 1,L1_1,TRUE\n"
    with pytest.raises(ValueError):
        parse_pest_file(_csv(tmp_path, "mixed.csv", bad))


# ---- service ---------------------------------------------------------------

def test_commit_status_and_export_block(isolated_db, tmp_path):
    weeks_service.create_week("2026-W22")
    path = _csv(tmp_path, "corn.csv", CORN_CSV)

    res = pests.commit(path, "2026-W22", 1)
    assert res.crop_code == "corn"
    assert res.imported == 3 and res.cards_completed == 3

    st = pests.pest_status("corn", "2026-W22")
    assert st["uploaded"] is True
    assert st["cards_completed"] == 3
    assert st["total_locations"] == 9
    assert set(st["bug_types"]) == {"Aphid", "Lygus", "fly spp."}

    bug_names, by_loc = pests.export_block("corn", "2026-W22")
    assert set(bug_names) == {"Aphid", "Lygus", "fly spp."}
    assert by_loc["L1"] == {"Lygus": "2", "fly spp.": "101"}
    assert "L4" not in by_loc  # not in the sheet

    # committing a week with no bugs (week 2) clears the block
    pests.commit(path, "2026-W22", 2)
    bug_names2, _ = pests.export_block("corn", "2026-W22")
    assert bug_names2 == []


# ---- export insertion ------------------------------------------------------

def test_export_inserts_bug_block_before_lab(isolated_db, corn, tmp_path):
    from app.exporters.excel_export import export_excel

    weeks_service.create_week("2026-W22")
    pests.commit(_csv(tmp_path, "corn.csv", CORN_CSV), "2026-W22", 1)

    out = tmp_path / "corn.xlsx"
    export_excel("corn", "2026-W22", out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]

    # bug columns are present...
    assert "Aphid" in headers and "fly spp." in headers
    # ...and sit before the first lab-page column (e.g. a nutrient/report col).
    fields = read_template_fields(corn.template_path)
    first_lab_name = min(
        (f for f in fields if f.name not in ("ID", "Location") and page_of(f) == "lab"),
        key=lambda f: f.excel_col,
    ).name
    assert headers.index("Aphid") < headers.index(first_lab_name)

    # L1's fly count landed on L1's row.
    id_col = headers.index("ID") + 1
    fly_col = headers.index("fly spp.") + 1
    l1_row = next(r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=id_col).value == "L1")
    assert ws.cell(row=l1_row, column=fly_col).value == 101
    wb.close()


def test_export_no_block_without_pest_data(isolated_db, corn, tmp_path):
    from app.exporters.excel_export import export_excel

    weeks_service.create_week("2026-W22")
    out = tmp_path / "corn.xlsx"
    export_excel("corn", "2026-W22", out)
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Aphid" not in headers  # no pest data -> unchanged export
    wb.close()


def test_weeks_progress_includes_pest_track(isolated_db, tmp_path):
    weeks_service.create_week("2026-W22")
    pests.commit(_csv(tmp_path, "corn.csv", CORN_CSV), "2026-W22", 1)  # 3 cards TRUE
    prog = weeks_service.all_weeks_progress()
    wk = next(w for w in prog if w["iso_week"] == "2026-W22")
    corn = next(c for c in wk["crops"] if c["crop_code"] == "corn")
    assert corn["pest_cards"] == 3
    assert corn["pest_uploaded"] is True


# ---- pest counts in Trends -------------------------------------------------

def test_pest_trends_category(isolated_db, tmp_path):
    from app.services import trends

    wk1 = (
        "p,DATE,ID Number,CARD COMPLETED,Aphid,Lygus\n"
        ",June 1,L1_1,TRUE,2,1\n"
        ",June 1,L2_1,TRUE,3,\n"
    )
    wk2 = (
        "p,DATE,ID Number,CARD COMPLETED,Aphid,Lygus\n"
        ",June 8,L1_1,TRUE,5,4\n"
    )
    weeks_service.create_week("2026-W22")
    weeks_service.create_week("2026-W23")
    pests.commit(_csv(tmp_path, "w1.csv", wk1), "2026-W22", 1)
    pests.commit(_csv(tmp_path, "w2.csv", wk2), "2026-W23", 1)

    field = trends.trend_series("corn", None, "pests")
    assert field["category"] == "pests"
    assert field["weeks"] == ["2026-W22", "2026-W23"]
    aphid = next(s for s in field["series"] if s["key"] == "Aphid")
    assert aphid["points"] == [5, 5]   # W22: 2+3, W23: 5

    l1 = trends.trend_series("corn", "L1", "pests")
    aphid_l1 = next(s for s in l1["series"] if s["key"] == "Aphid")
    assert aphid_l1["points"] == [2, 5]


# ---- rename migrates pest rows --------------------------------------------

def test_rename_week_migrates_pest(isolated_db, tmp_path):
    weeks_service.create_week("2026-W22")
    pests.commit(_csv(tmp_path, "corn.csv", CORN_CSV), "2026-W22", 1)
    weeks_service.rename_week("2026-W22", "Scout-1")
    assert pests.pest_status("corn", "Scout-1")["uploaded"] is True
    assert pests.pest_status("corn", "2026-W22")["uploaded"] is False
