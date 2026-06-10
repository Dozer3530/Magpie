"""Web-frontend parity tests.

The FastAPI layer must be a faithful, thin shell over the same services the
desktop uses. These tests drive the HTTP API end-to-end and assert:

  * the read routes return what the services compute, and
  * a week built *through the API* produces the same export files the desktop
    golden test produces for the same inputs (the parity guarantee).

Isolation reuses the `isolated_db` fixture; we additionally redirect the
exporter's EXPORTS_DIR into tmp so built packages land under tmp_path.
"""
from __future__ import annotations

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.schema import read_template_fields
from app.services import exports as exports_service


@pytest.fixture
def client(isolated_db, monkeypatch):
    # Built packages go under tmp so we don't touch the real exports/ folder.
    monkeypatch.setattr(exports_service, "EXPORTS_DIR", isolated_db / "exports")
    # Import inside the fixture so the patched globals are already in place
    # when the app's startup event runs init_db().
    from webapp.server import app

    with TestClient(app) as c:
        yield c


def _field_name(crop_template, key):
    for f in read_template_fields(crop_template):
        if f.key == key:
            return f.name
    raise KeyError(key)


# ---- Read routes -----------------------------------------------------------

def test_crops_and_weeks(client):
    crops = client.get("/api/crops").json()
    assert {c["code"] for c in crops} == {"canola", "corn"}

    # startup auto-seeds the current week.
    weeks = client.get("/api/weeks").json()
    assert weeks["current"]
    assert len(weeks["weeks"]) >= 1


def test_form_schema_section_order(client):
    schema = client.get("/api/form-schema", params={"crop": "canola"}).json()
    titles = [s["title"] for s in schema["sections"]]
    assert titles == [
        "Observation header",
        "Soil readings (TDR)",
        "Diseases",
        "Insects",
        "Petal test",
        "Photos",
        "Notes",
        "Report identifiers",
        "Nutrient panel",
        "Nutrient ratios",
    ]
    # Corn has no petal test.
    corn = client.get("/api/form-schema", params={"crop": "corn"}).json()
    assert "Petal test" not in [s["title"] for s in corn["sections"]]


def test_week_crud_and_obs_roundtrip(client, canola):
    week = "2026-W22"
    assert client.post("/api/weeks", json={"tag": week}).json()["ok"]

    n_name = _field_name(canola.template_path, "N")
    save = client.put("/api/obs", json={
        "crop": "canola", "week": week, "loc": "M1",
        "values": {n_name: "3.2", "Disease_Blackleg": "yes"},
    })
    assert save.json()["ok"]

    got = client.get("/api/obs", params={"crop": "canola", "week": week, "loc": "M1"}).json()
    assert got["values"][n_name] == "3.2"
    assert got["values"]["Disease_Blackleg"] == "yes"

    overview = client.get("/api/overview", params={"crop": "canola", "week": week}).json()
    assert overview["total_locations"] == 9
    assert overview["locations_with_data"] == 1


# ---- Weeks: progress dashboard + rename ------------------------------------

def test_weeks_progress_two_tracks(client, canola):
    week = "2026-W22"
    client.post("/api/weeks", json={"tag": week})
    n_name = _field_name(canola.template_path, "N")          # lab column
    client.put("/api/obs", json={"crop": "canola", "week": week, "loc": "M1",
                                 "values": {"Disease_Blackleg": "yes"}})   # field
    client.put("/api/obs", json={"crop": "canola", "week": week, "loc": "M2",
                                 "values": {n_name: "3.0"}})               # lab

    prog = client.get("/api/weeks/progress").json()
    wk = next(w for w in prog if w["iso_week"] == week)
    canola_p = next(c for c in wk["crops"] if c["crop_code"] == "canola")
    assert canola_p["field_locations"] == 1
    assert canola_p["lab_locations"] == 1


def test_week_rename_route(client, canola):
    week = "2026-W22"
    client.post("/api/weeks", json={"tag": week})
    n_name = _field_name(canola.template_path, "N")
    client.put("/api/obs", json={"crop": "canola", "week": week, "loc": "M1",
                                 "values": {n_name: "3.2"}})

    res = client.post("/api/weeks/rename", json={"old": week, "new": "Bloom-1"})
    assert res.status_code == 200 and res.json()["tag"] == "Bloom-1"

    tags = [w["iso_week"] for w in client.get("/api/weeks").json()["weeks"]]
    assert "Bloom-1" in tags and week not in tags
    got = client.get("/api/obs", params={"crop": "canola", "week": "Bloom-1", "loc": "M1"}).json()
    assert got["values"][n_name] == "3.2"


def test_week_rename_bad_name_400(client):
    week = "2026-W22"
    client.post("/api/weeks", json={"tag": week})
    res = client.post("/api/weeks/rename", json={"old": week, "new": "bad/name"})
    assert res.status_code == 400


def test_trends_route(client, canola):
    t1 = _field_name(canola.template_path, "TDR_1_SOIL_TEMPERATURE")
    client.post("/api/weeks", json={"tag": "2026-W22"})
    client.put("/api/obs", json={"crop": "canola", "week": "2026-W22", "loc": "M1",
                                 "values": {t1: "17.5"}})
    field = client.get("/api/trends", params={"crop": "canola", "cat": "soil"}).json()
    assert "2026-W22" in field["weeks"]
    temp = next(s for s in field["series"] if s["key"] == "temp")
    assert temp["points"][field["weeks"].index("2026-W22")] == 17.5
    # per-point scope + a different category
    m1 = client.get("/api/trends", params={"crop": "canola", "loc": "M1", "cat": "nutrients"}).json()
    assert m1["scope"] == "M1" and m1["category"] == "nutrients"


def test_backup_route(client, isolated_db):
    res = client.post("/api/backup")
    assert res.status_code == 200
    body = res.json()
    assert body["ok"] and body["name"].startswith("packages_")
    # the backup file actually exists under the isolated data dir
    assert (isolated_db / "backups" / body["name"]).exists()


def test_publish_routes(client, tmp_path, monkeypatch):
    from app import app_settings
    monkeypatch.setattr(app_settings, "_SETTINGS_PATH", tmp_path / "settings.json")
    drive = tmp_path / "drive"
    assert client.put("/api/publish/dir", json={"path": str(drive)}).json()["ok"]
    client.post("/api/weeks", json={"tag": "2026-W22"})
    res = client.post("/api/publish").json()
    assert res["ok"]
    assert (drive / "magpie-progress.html").exists()


def test_publish_no_dir_400(client, tmp_path, monkeypatch):
    from app import app_settings
    monkeypatch.setattr(app_settings, "_SETTINGS_PATH", tmp_path / "settings2.json")
    assert client.post("/api/publish").status_code == 400


# ---- Import flow (upload -> commit) ----------------------------------------

def test_survey_import_upload_then_commit(client, canola):
    week = "2026-W23"
    client.post("/api/weeks", json={"tag": week})
    n_name = _field_name(canola.template_path, "N")

    csv = f"ID,N\nM1,2.5\nM2,4.1\nZZ,9.9\n".encode()
    up = client.post(
        "/api/import/upload",
        data={"crop": "canola"},
        files={"file": ("survey.csv", io.BytesIO(csv), "text/csv")},
    ).json()
    token = up["token"]
    assert up["row_count"] == 3
    # "N" auto-maps to the unit-bearing template header.
    assert up["mapping"]["matches"]["N"] == n_name

    res = client.post("/api/import/survey", json={
        "token": token, "crop": "canola", "week": week, "id_col": "ID",
    }).json()
    assert res["imported"] == 2          # M1, M2
    assert res["skipped_no_loc"] == 1    # ZZ

    got = client.get("/api/obs", params={"crop": "canola", "week": week, "loc": "M2"}).json()
    assert got["values"][n_name] == "4.1"


def test_scouting_upload_and_commit(client, canola):
    from app.schema import read_template_locations
    loc_id, lat, lon = read_template_locations(canola.template_path)[0]
    week = "2026-W24"
    client.post("/api/weeks", json={"tag": week})
    csv = (
        "Date &  Time,Canola or Corn?,Canola Crop Growth Stage,Scouters Name,x,y\n"
        f'6/9/2026 4:05:00 PM,Canola,10 - Cotyledons completely unfold,Christina,{lon},{lat}\n'
    ).encode()
    up = client.post("/api/scouting/upload",
                     files={"file": ("scout.csv", io.BytesIO(csv), "text/csv")}).json()
    assert up["events"][0]["date"] == "2026-06-09"
    assert up["events"][0]["matched"] == 1
    a = up["events"][0]["assignments"][0]
    assert a["point"] == loc_id and a["status"] == "matched"

    res = client.post("/api/scouting/commit",
                      json={"token": up["token"], "week": week, "date": "2026-06-09"}).json()
    assert res["imported"] == {"canola": 1}

    got = client.get("/api/obs", params={"crop": "canola", "week": week, "loc": loc_id}).json()
    assert got["values"]["Canola_Crop_Growth_Stage"] == "10 - Cotyledons completely unfold"


def test_pest_upload_commit_status(client, corn):
    week = "2026-W22"
    client.post("/api/weeks", json={"tag": week})
    csv = (
        b"person,DATE,ID Number,CARD COMPLETED,Aphid,fly spp.\n"
        b",June 1,L1_1,TRUE,2,101\n"
        b",June 1,L2_1,TRUE,,55\n"
        b",June 8,L1_2,FALSE,,\n"
    )
    up = client.post("/api/pest/upload",
                     files={"file": ("pest.csv", io.BytesIO(csv), "text/csv")}).json()
    assert up["crop"] == "corn"
    assert [w["index"] for w in up["weeks"]] == [1, 2]

    res = client.post("/api/pest/commit",
                      json={"token": up["token"], "week": week, "week_index": 1}).json()
    assert res["ok"] and res["cards_completed"] == 2
    assert set(res["bug_types"]) == {"Aphid", "fly spp."}

    st = client.get("/api/pest/status", params={"crop": "corn", "week": week}).json()
    assert st["uploaded"] is True and st["total_locations"] == 9


def test_lab_import_duplicate_target_conflicts(client):
    week = "2026-W24"
    client.post("/api/weeks", json={"tag": week})

    csv = b"SampleID,N\nPoint A,1.1\nPoint B,2.2\n"
    up = client.post(
        "/api/import/upload",
        data={"crop": "canola"},
        files={"file": ("lab.csv", io.BytesIO(csv), "text/csv")},
    ).json()

    # Assign both rows to M1 → must 409.
    resp = client.post("/api/import/lab", json={
        "token": up["token"], "crop": "canola", "week": week,
        "row_targets": {"0": "M1", "1": "M1"},
    })
    assert resp.status_code == 409
    assert "M1" in resp.json()["detail"]


# ---- Export parity ---------------------------------------------------------

def test_export_through_api_matches_golden(client, canola):
    """Build a week via the API; assert the xlsx matches the golden shape."""
    week = "2026-W25"
    client.post("/api/weeks", json={"tag": week})
    n_name = _field_name(canola.template_path, "N")
    client.put("/api/obs", json={
        "crop": "canola", "week": week, "loc": "M1",
        "values": {n_name: "3.2", "Disease_Blackleg": "yes"},
    })

    res = client.post("/api/export", params={"crop": "canola", "week": week}).json()
    assert res["zip_name"] == f"EarthDaily_{week}.zip"
    assert not res["errors"]

    # The download route serves the built zip.
    dl = client.get(f"/api/export/{week}/download")
    assert dl.status_code == 200
    assert dl.headers["content-type"] == "application/zip"

    # Same assertions as the desktop golden export, on the API-built xlsx.
    from app.exporters.excel_export import export_filename
    xlsx = exports_service.week_dir(week) / export_filename("canola", week)
    fields = read_template_fields(canola.template_path)
    wb = openpyxl.load_workbook(xlsx)
    ws = wb.active
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(fields) + 1)]
    assert headers == [f.name for f in fields]
    col = {f.name: f.excel_col for f in fields}
    assert ws.cell(row=2, column=col[n_name]).value == 3.2
    assert ws.cell(row=2, column=col["Disease_Blackleg"]).value == "yes"
    wb.close()
