"""Tests for the Reactive feed (client-scattered, non-home points)."""
from __future__ import annotations

import zipfile

import openpyxl
import pytest

from app import db
from app.schema import read_template_locations
from app.services import exports
from app.services import reactive
from app.services import scouting
from app.services import weeks as weeks_service

HDR = (
    "Date &  Time,Where are you?,Canola or Corn?,"
    "Canola Crop Growth Stage,Corn Crop Growth Stage,#1 TDR - SOIL MOISTURE,"
    "Please name the insect(s) ,Insect Damage Severity,Optional Additional Notes,x,y"
)


def _row(when, field, crop, lat, lon, **kw):
    cols = [
        when, field, crop,
        kw.get("canola_gs", ""), kw.get("corn_gs", ""), kw.get("moist", ""),
        kw.get("insect", ""), kw.get("sev", ""), kw.get("notes", ""),
        str(lon), str(lat),
    ]
    return ",".join(cols)


def _write(tmp_path, rows, name="scout.csv"):
    p = tmp_path / name
    p.write_text(HDR + "\n" + "\n".join(rows) + "\n", encoding="utf-8")
    return p


def _loc(crop, idx):
    return read_template_locations(crop.template_path)[idx]  # (id, lat, lon)


def test_partition_home_stays_in_scouting_reactive_only_others(isolated_db, tmp_path, canola):
    m1 = _loc(canola, 0)
    p = _write(tmp_path, [
        # home field — scouting owns it (sits on the M1 stake)
        _row("6/9/2026 16:05", "Field 18", "Canola", m1[1], m1[2], canola_gs="10 - x", moist="27"),
        # reactive field — its own GPS, far from any stake
        _row("6/9/2026 16:20", "Field F", "Canola", 51.60, -114.20, canola_gs="09 - x", moist="31"),
    ])

    rev = reactive.prepare(p)["events"][0]
    assert rev["n_rows"] == 1
    assert rev["assignments"][0]["field"] == "Field F"
    assert rev["assignments"][0]["point"] == "F1"

    sev = scouting.prepare(p)["events"][0]
    by_field = {a["field"]: a for a in sev["assignments"]}
    assert by_field["Field 18"]["status"] == "matched" and by_field["Field 18"]["point"] == m1[0]
    assert by_field["Field F"]["status"] == "other_field"  # never GPS-matched here


def test_numbering_continues_across_weeks_and_is_idempotent(isolated_db, tmp_path):
    weeks_service.create_week("2026-W24")
    weeks_service.create_week("2026-W26")
    wk1 = _write(tmp_path, [
        _row("6/9/2026 16:20", "Field F", "Canola", 51.60, -114.20, canola_gs="10 - x"),
        _row("6/9/2026 16:35", "Field F", "Canola", 51.61, -114.21, canola_gs="10 - x"),
        _row("6/9/2026 16:50", "Field F", "Canola", 51.62, -114.22, canola_gs="10 - x"),
    ], "wk1.csv")
    wk3 = _write(tmp_path, [
        _row("6/23/2026 15:00", "Field F", "Canola", 51.63, -114.23, canola_gs="12 - x"),
        _row("6/23/2026 15:10", "Field F", "Canola", 51.64, -114.24, canola_gs="12 - x"),
    ], "wk3.csv")

    reactive.commit(wk1, "2026-W24", "2026-06-09")
    reactive.commit(wk3, "2026-W26", "2026-06-23")
    assert [d["point_id"] for d in reactive.week_points("canola", "2026-W24")] == ["F1", "F2", "F3"]
    assert [d["point_id"] for d in reactive.week_points("canola", "2026-W26")] == ["F4", "F5"]

    # Re-importing week 1 must not renumber or duplicate.
    reactive.commit(wk1, "2026-W24", "2026-06-09")
    assert [d["point_id"] for d in reactive.week_points("canola", "2026-W24")] == ["F1", "F2", "F3"]
    with db.connect() as conn:
        assert conn.execute("SELECT COUNT(*) FROM reactive_obs").fetchone()[0] == 5


def test_commit_translates_both_crops(isolated_db, tmp_path):
    weeks_service.create_week("2026-W24")
    p = _write(tmp_path, [
        _row("6/9/2026 16:20", "Field F", "Canola", 51.60, -114.20,
             canola_gs="09 - Cotyledons", moist="31", insect="Flea beetle", sev="Low", notes="scattered A"),
        _row("6/9/2026 17:00", "Field G", "Corn", 51.55, -114.30,
             corn_gs="10 - First leaf", moist="33", sev="Medium", notes="corn scatter"),
    ])
    res = reactive.commit(p, "2026-W24", "2026-06-09")
    assert res["imported"] == {"canola": 1, "corn": 1}
    assert res["unmapped_columns"] == []
    assert set(res["fields"]) == {"Field F", "Field G"}

    f1 = reactive.week_points("canola", "2026-W24")[0]
    assert f1["point_id"] == "F1" and f1["field"] == "Field F"
    assert f1["values"]["Canola_Crop_Growth_Stage"] == "09 - Cotyledons"
    assert f1["values"]["Insect_Identification"] == "Flea beetle"
    assert f1["values"]["Notes"] == "scattered A"

    g1 = reactive.week_points("corn", "2026-W24")[0]
    assert g1["point_id"] == "G1"
    assert g1["values"]["Insect_Damage_Severity"] == "Med"  # Medium -> Med


def test_export_reactive_files_in_zip_without_lab(isolated_db, tmp_path, monkeypatch):
    monkeypatch.setattr(exports, "EXPORTS_DIR", tmp_path / "exports")
    weeks_service.create_week("2026-W24")
    p = _write(tmp_path, [
        _row("6/9/2026 16:20", "Field F", "Canola", 51.60, -114.20, canola_gs="10 - x", moist="31"),
        _row("6/9/2026 17:00", "Field G", "Corn", 51.55, -114.30, corn_gs="10 - y", moist="33"),
    ])
    reactive.commit(p, "2026-W24", "2026-06-09")

    res = exports.build_all("2026-W24")
    assert not res.errors
    names = {pth.name for pth in res.produced}
    assert "Reactive_Canola_2026-W24.xlsx" in names
    assert "Reactive_Corn_2026-W24.gpkg" in names

    zpath = tmp_path / "exports" / "2026-W24" / "EarthDaily_2026-W24.zip"
    with zipfile.ZipFile(zpath) as z:
        assert "Reactive_Canola_2026-W24.xlsx" in z.namelist()

    wb = openpyxl.load_workbook(tmp_path / "exports" / "2026-W24" / "Reactive_Canola_2026-W24.xlsx")
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert headers[:3] == ["ID", "Field", "Location"]
    assert "N (%)" not in headers and "N" not in headers   # no lab columns
    assert ws.cell(2, 1).value == "F1"
    assert ws.max_row == 2  # one canola reactive point


def test_no_reactive_data_means_no_reactive_files(isolated_db, tmp_path, monkeypatch, canola):
    """A week with only fixed-point data must produce an unchanged package."""
    monkeypatch.setattr(exports, "EXPORTS_DIR", tmp_path / "exports")
    weeks_service.create_week("2026-W24")
    # Commit one ordinary fixed-stake scouting row (home field).
    m1 = _loc(canola, 0)
    p = _write(tmp_path, [
        _row("6/9/2026 16:05", "Field 18", "Canola", m1[1], m1[2], canola_gs="10 - x", moist="27"),
    ])
    scouting.commit(p, "2026-W24", "2026-06-09")

    res = exports.build_all("2026-W24")
    assert not any("Reactive_" in pth.name for pth in res.produced)


def test_rename_week_migrates_reactive(isolated_db, tmp_path):
    weeks_service.create_week("2026-W24")
    p = _write(tmp_path, [
        _row("6/9/2026 16:20", "Field F", "Canola", 51.60, -114.20, canola_gs="10 - x"),
    ])
    reactive.commit(p, "2026-W24", "2026-06-09")
    weeks_service.rename_week("2026-W24", "2026-W25")
    assert reactive.week_points("canola", "2026-W24") == []
    assert [d["point_id"] for d in reactive.week_points("canola", "2026-W25")] == ["F1"]
