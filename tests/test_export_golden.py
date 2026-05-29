"""Golden end-to-end export test — the regression oracle for the refactor.

Seed a known week, run the real Excel + GPKG exporters, and assert the
produced files match the template's shape with stamped values. The upcoming
service-layer extraction must keep these outputs identical.
"""
from __future__ import annotations

import geopandas as gpd
import openpyxl

from app import db
from app.exporters.excel_export import export_excel
from app.exporters.gpkg_export import export_gpkg
from app.schema import read_template_fields
from tests.conftest import seed_week


def _field_by_key(fields, key):
    for f in fields:
        if f.key == key:
            return f
    raise KeyError(key)


def test_golden_excel_export(isolated_db, canola):
    week = seed_week()
    fields = read_template_fields(canola.template_path)
    n_field = _field_by_key(fields, "N")

    with db.connect() as conn:
        db.save_obs_row(conn, "canola", week, "M1", {
            n_field.name: "3.2",
            "Disease_Blackleg": "yes",
        })

    out = isolated_db / "Canola.xlsx"
    export_excel("canola", week, out)
    assert out.is_file()

    wb = openpyxl.load_workbook(out)
    ws = wb.active

    # Row-1 headers match the template field list exactly (order + text).
    headers = [ws.cell(row=1, column=c).value for c in range(1, len(fields) + 1)]
    assert headers == [f.name for f in fields]

    # 9 location rows (M1..M9), in order.
    ids = [ws.cell(row=r, column=1).value for r in range(2, 11)]
    assert ids == [f"M{i}" for i in range(1, 10)]

    # M1's stamped values: N coerced numeric, disease presence as text.
    col_by_name = {f.name: f.excel_col for f in fields}
    assert ws.cell(row=2, column=col_by_name[n_field.name]).value == 3.2
    assert ws.cell(row=2, column=col_by_name["Disease_Blackleg"]).value == "yes"
    # An unfilled location stays blank.
    assert ws.cell(row=10, column=col_by_name["Disease_Blackleg"]).value is None
    wb.close()


def test_golden_gpkg_export(isolated_db, canola):
    week = seed_week()
    fields = read_template_fields(canola.template_path)
    n_field = _field_by_key(fields, "N")

    with db.connect() as conn:
        db.save_obs_row(conn, "canola", week, "M1", {n_field.name: "3.2"})

    out = isolated_db / "Canola.gpkg"
    export_gpkg("canola", week, out)
    assert out.is_file()

    gdf = gpd.read_file(out)
    # One point feature per location.
    assert len(gdf) == 9
    assert gdf.crs is not None and gdf.crs.to_epsg() == 4326
    assert (gdf.geometry.geom_type == "Point").all()

    m1 = gdf[gdf["ID"] == "M1"].iloc[0]
    assert float(m1[n_field.name]) == 3.2
