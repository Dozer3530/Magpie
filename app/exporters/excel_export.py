"""Stamp the weekly observation rows into a copy of the crop template.

The export is a copy of the per-crop `Static <Crop> Template.xlsx` with rows
2..N filled in from the DB. We preserve the template's styling (fonts,
column widths, header formatting) by writing into the loaded workbook and
saving it under the new path; we never re-create the file from scratch.

One row per location, in `locations` table order (M1..M9 / L1..L9). Locations
with no observation for the week get an ID + Location entry and blank cells
everywhere else.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app import image_storage
from app.crops import CropConfig, crop_by_code
from app.db import connect, list_locations, list_obs_for_week
from app.schema import Field, FieldKind, page_of, read_template_fields


def _coerce_value(field: Field, raw: str):
    """Convert a stored TEXT value to the right Python type for openpyxl.

    Blank → None (empty cell). NUMBER → float (or int if it round-trips).
    Everything else → str.
    """
    if raw == "" or raw is None:
        return None
    if field.kind == FieldKind.NUMBER:
        try:
            f = float(raw)
            return int(f) if f.is_integer() else f
        except ValueError:
            return raw
    return raw


def _images_formula(crop_code: str, location_id: str, raw: str) -> str | None:
    """Build the HYPERLINK formula for the Images cell.

    - 1 photo → link directly to the file
    - 2+ photos → link to the location's folder so the client sees them all
    """
    names = image_storage.parse_list(raw)
    if not names:
        return None
    rel_base = f"images/{crop_code}/{location_id}/"
    if len(names) == 1:
        # Excel formulas use "" to escape inner quotes; our paths/names don't
        # contain quotes so a plain interpolation is safe.
        return f'=HYPERLINK("{rel_base}{names[0]}","{names[0]}")'
    return f'=HYPERLINK("{rel_base}","{len(names)} photos")'


def export_excel(crop_code: str, iso_week: str, out_path: Path) -> Path:
    """Write `<Crop>_<iso_week>.xlsx` to `out_path`. Returns the written path."""
    crop: CropConfig = crop_by_code(crop_code)
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(crop.template_path)
    ws = wb.active

    fields = read_template_fields(crop.template_path)
    col_by_name = {f.name: f.excel_col for f in fields}
    field_by_name = {f.name: f for f in fields}

    with connect() as conn:
        locations = list_locations(conn, crop_code)
        obs_rows = list_obs_for_week(conn, crop_code, iso_week)
    obs_by_loc = {row["location_id"]: dict(row) for row in obs_rows}

    # Clear any pre-existing data rows beyond row 1 in the template (the
    # shipped templates have empty A2..N rows with ID/Location pre-filled —
    # we'll rewrite them all so the template's lat/long is authoritative).
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    loc_row: dict[str, int] = {}
    for r, loc in enumerate(locations, start=2):
        loc_id = loc["location_id"]
        loc_row[loc_id] = r
        if "ID" in col_by_name:
            ws.cell(row=r, column=col_by_name["ID"], value=loc_id)
        if "Location" in col_by_name:
            ws.cell(
                row=r,
                column=col_by_name["Location"],
                value=f"{loc['lat']}, {loc['lon']}",
            )
        obs = obs_by_loc.get(loc_id, {})
        for name, col in col_by_name.items():
            if name in ("ID", "Location"):
                continue
            field = field_by_name[name]
            if field.kind == FieldKind.IMAGES:
                formula = _images_formula(crop_code, loc_id, obs.get(name, ""))
                if formula is not None:
                    ws.cell(row=r, column=col, value=formula)
                continue
            value = _coerce_value(field, obs.get(name, ""))
            if value is not None:
                ws.cell(row=r, column=col, value=value)

    _write_pest_block(ws, fields, crop_code, iso_week, loc_row)
    _autofit_columns(ws)

    wb.save(out_path)
    return out_path


# Header fill for the inserted pest block — a green nod to the sheet's
# "CARD COMPLETED GREEN" cue, distinct from the rest of the template.
_PEST_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_PEST_FONT = Font(bold=True)


def _pest_value(raw: str):
    """Counts are integers in the sheet; write them as numbers when they parse."""
    try:
        f = float(raw)
        return int(f) if f.is_integer() else f
    except (TypeError, ValueError):
        return raw


def _write_pest_block(ws, fields, crop_code: str, iso_week: str, loc_row: dict[str, int]) -> None:
    """Insert a colored block of the week's bug-count columns before the lab
    nutrients. No-op when there's no pest data for the week (so the normal
    export is byte-for-byte unchanged). Falls back to appending the block at
    the end if mid-sheet insertion isn't possible on this template.
    """
    from app.services import pests  # lazy: avoids any import cycle

    bug_names, by_loc = pests.export_block(crop_code, iso_week)
    if not bug_names:
        return

    lab_cols = [
        f.excel_col for f in fields
        if f.name not in ("ID", "Location") and page_of(f) == "lab"
    ]
    n = len(bug_names)
    try:
        insert_at = min(lab_cols) if lab_cols else (ws.max_column + 1)
        if insert_at <= ws.max_column:
            ws.insert_cols(insert_at, n)
        start = insert_at
    except Exception:
        # Insertion not possible (e.g. merged cells) — append at the far right.
        start = ws.max_column + 1

    for i, bug in enumerate(bug_names):
        c = ws.cell(row=1, column=start + i, value=bug)
        c.fill = _PEST_FILL
        c.font = _PEST_FONT
    for loc_id, bugs in by_loc.items():
        r = loc_row.get(loc_id)
        if r is None:
            continue
        for i, bug in enumerate(bug_names):
            if bug in bugs:
                ws.cell(row=r, column=start + i, value=_pest_value(bugs[bug]))


def _autofit_columns(ws, max_width: int = 48, padding: int = 2) -> None:
    """Size every column to its widest cell — what you'd get by selecting all
    columns and double-clicking a divider in Excel — so long headers like
    `TDR_1_SOIL_TEMPERATURE (°C)` aren't smooshed on open. Width is capped so a
    stray long value can't blow a column out to the whole screen.
    """
    for col in range(1, ws.max_column + 1):
        longest = 0
        for row in range(1, ws.max_row + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                continue
            text = str(value)
            if text.startswith("="):
                # HYPERLINK formula: measure the visible label, not the formula.
                # =HYPERLINK("target","LABEL")  ->  LABEL
                if '","' in text:
                    text = text.rsplit('","', 1)[-1].rstrip('")')
            longest = max(longest, len(text))
        if longest:
            ws.column_dimensions[get_column_letter(col)].width = min(longest + padding, max_width)


def export_filename(crop_code: str, iso_week: str) -> str:
    crop = crop_by_code(crop_code)
    return f"{crop.display_name}_{iso_week}.xlsx"
