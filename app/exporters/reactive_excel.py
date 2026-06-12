"""Stamp a week's reactive points into a copy of the Reactive <Crop> template.

Parallels `excel_export.py` but the rows come from `reactive_obs` (via
`reactive.week_points`) instead of the fixed `locations` table: each point
supplies its own derived ID (F1, F2...), field name, and "lat, lon". No lab
columns, no pest block, no photos — reactive is Survey123-only.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl

from app.crops import crop_by_code
from app.exporters.excel_export import _autofit_columns, _coerce_value
from app.schema import FieldKind, read_template_fields
from app.services import reactive as reactive_service

_DIRECT = ("ID", "Field", "Location")  # filled from the point, not the survey row


def export_reactive_excel(crop_code: str, iso_week: str, out_path: Path) -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    template = reactive_service.reactive_template_path(crop_code)
    wb = openpyxl.load_workbook(template)
    ws = wb.active

    fields = read_template_fields(template)
    col_by_name = {f.name: f.excel_col for f in fields}
    field_by_name = {f.name: f for f in fields}
    points = reactive_service.week_points(crop_code, iso_week)

    # Clear the template's (empty) data rows, then write one row per point.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for r, pt in enumerate(points, start=2):
        if "ID" in col_by_name:
            ws.cell(row=r, column=col_by_name["ID"], value=pt["point_id"])
        if "Field" in col_by_name:
            ws.cell(row=r, column=col_by_name["Field"], value=pt["field"])
        if "Location" in col_by_name:
            ws.cell(row=r, column=col_by_name["Location"], value=f"{pt['lat']}, {pt['lon']}")
        for name, col in col_by_name.items():
            if name in _DIRECT:
                continue
            field = field_by_name[name]
            if field.kind == FieldKind.IMAGES:
                continue  # reactive points carry no photos
            value = _coerce_value(field, pt["values"].get(name, ""))
            if value is not None:
                ws.cell(row=r, column=col, value=value)

    _autofit_columns(ws)
    wb.save(out_path)
    return out_path


def export_filename(crop_code: str, iso_week: str) -> str:
    return f"Reactive_{crop_by_code(crop_code).display_name}_{iso_week}.xlsx"
