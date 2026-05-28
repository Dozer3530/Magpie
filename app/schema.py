"""Read template headers, classify columns into typed fields.

Each template (`Static <Crop> Template.xlsx`) is the source of truth for the
column list of its `obs_<crop>` table. This module reads row 1 of a template
and returns a list of typed `Field` records that downstream code uses to:

  - generate the SQLite table schema (`db.py`)
  - lay out the per-location form widgets (`ui/observations_tab.py`)
  - validate imported lab/Survey123 values
  - write the weekly Excel export back into the same column order

The classification is *pattern-based* on column names — adding a new disease
or nutrient to a template is automatically picked up without code changes.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter


# ---- Field taxonomy ---------------------------------------------------------

class FieldKind(str, Enum):
    KEY = "key"                  # ID (row identifier from template col A)
    LOCATION = "location"        # "lat, lon" string from template col B
    DATETIME = "datetime"        # observation timestamp
    GROWTH_STAGE = "growth_stage"  # picker from per-crop BBCH list
    DISEASE_PRESENCE = "disease_presence"  # "yes" or blank
    SEVERITY = "severity"        # Low / Med / High or blank
    NUMBER = "number"            # numeric reading (TDR, nutrient value, ratio, %)
    RATING = "rating"            # nutrient `_rate` letter code (D/L/S/H/VH) or blank
    TEXT = "text"                # free-form text (insect description, lab no., report no.)
    IMAGES = "images"            # filepath(s); skipped in v1


SEVERITY_CHOICES = ("", "Low", "Med", "High")
PRESENCE_CHOICES = ("", "yes")
# Standard PT2R nutrient sufficiency codes. The combo is editable so any
# one-off lab code (e.g. "M" for medium) still round-trips through the UI.
RATING_CHOICES = ("", "D", "L", "S", "H", "VH")

# Display units baked into the template headers (e.g. `N (%)`, `TDR_1_SOIL_EC
# (dS/m)`). The unit is part of the header text the client sees, but internal
# logic (classification, pairing, import matching) works on the unit-stripped
# *key* (see `_strip_unit` / `Field.key`). This table documents the units we
# stamp; the templates themselves remain the source of truth.
#
# NOTE: NO3N, Na, and Cl vary by lab (% vs ppm). These reflect the common
# Canadian plant-tissue convention — verify against the client's PT2R report.
UNIT_BY_KEY = {
    "TDR_1_SOIL_TEMPERATURE": "°C", "TDR_2_SOIL_TEMPERATURE": "°C", "TDR_3_SOIL_TEMPERATURE": "°C",
    "TDR_1_SOIL_EC": "dS/m", "TDR_2_SOIL_EC": "dS/m", "TDR_3_SOIL_EC": "dS/m",
    "TDR_1_SOIL_MOISTURE": "%", "TDR_2_SOIL_MOISTURE": "%", "TDR_3_SOIL_MOISTURE": "%",
    "N": "%", "P": "%", "K": "%", "Ca": "%", "Mg": "%", "S": "%", "Na": "%", "Cl": "%",
    "NO3N": "ppm", "Zn": "ppm", "Mn": "ppm", "Fe": "ppm", "Cu": "ppm",
    "B": "ppm", "Mo": "ppm", "Al": "ppm",
}

# Matches a trailing " (unit)" parenthetical so the rest of the codebase can
# work on the canonical, unit-free column name.
_UNIT_SUFFIX_RE = re.compile(r"\s*\([^)]*\)\s*$")


def _strip_unit(name: str) -> str:
    """Return the canonical column key: header text minus a trailing ` (unit)`."""
    return _UNIT_SUFFIX_RE.sub("", str(name)).strip()


# Column-name patterns. Order matters: first match wins. Operates on the
# unit-stripped *key* (callers pass `_strip_unit(header)`).
# Severity (any `*_Severity`) is checked before disease presence.
def classify(header: str) -> FieldKind:
    h = header.strip()
    low = h.lower()

    if h == "ID":
        return FieldKind.KEY
    if h == "Location":
        return FieldKind.LOCATION
    if h == "Date_Time":
        return FieldKind.DATETIME
    if low.endswith("crop_growth_stage"):
        return FieldKind.GROWTH_STAGE
    if h == "Images":
        return FieldKind.IMAGES

    if low.endswith("_severity"):
        # Disease_*_Severity and Insect_Damage_Severity → Low/Med/High picker.
        return FieldKind.SEVERITY
    if low.startswith("disease_") and h != "Disease_Report_Results":
        # Disease_Report_Results (corn template) is free text, not yes/blank.
        return FieldKind.DISEASE_PRESENCE

    # Numeric readings: TDR sensors, lab nutrient values + rates,
    # petal-test counts and percent infected, all *_Actual / *_Expected ratios.
    if low.startswith("tdr_"):
        return FieldKind.NUMBER
    if low.startswith("petal_test_") and low != "petal_test_no.":
        return FieldKind.NUMBER
    if low.endswith("_actual") or low.endswith("_expected"):
        return FieldKind.NUMBER
    if low.endswith("_rate"):
        # Nutrient sufficiency letter code (PT2R): D/L/S/H/VH or blank.
        return FieldKind.RATING

    # Bare-element nutrient names (single or two-letter element symbols and NO3N).
    NUTRIENTS = {
        "N", "NO3N", "P", "K", "Ca", "Mg", "S",
        "Zn", "Mn", "Fe", "Cu", "B", "Mo", "Al", "Na", "Cl",
    }
    if h in NUTRIENTS:
        return FieldKind.NUMBER

    # Petal_Test_No., Lab_No., ReportNo, Insect_Damage, Insect_Identification,
    # Disease_Report_Results → text.
    return FieldKind.TEXT


# ---- Field record -----------------------------------------------------------

@dataclass(frozen=True)
class Field:
    name: str               # exact header text (used as DB column name)
    kind: FieldKind
    excel_col: int          # 1-based column index in the template
    choices: tuple[str, ...] = ()  # populated for SEVERITY / DISEASE_PRESENCE

    @property
    def excel_col_letter(self) -> str:
        return get_column_letter(self.excel_col)

    @property
    def key(self) -> str:
        """Canonical, unit-free column name (e.g. `N (%)` → `N`).

        All internal matching/pairing uses the key so units in the header
        don't break classification, nutrient/severity pairing, or imports.
        """
        return _strip_unit(self.name)

    @property
    def sql_type(self) -> str:
        # We store everything as TEXT so values round-trip the templates
        # byte-equivalent (e.g. "yes" presence, "Low" severity, blank cells).
        # Numeric fields are still TEXT in the DB but validated as numbers
        # in the UI layer.
        return "TEXT"


def read_template_fields(template_path: Path) -> list[Field]:
    """Return the typed Field list for a template's row-1 headers."""
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ws = wb.active
    fields: list[Field] = []
    for col in range(1, ws.max_column + 1):
        name = ws.cell(row=1, column=col).value
        if not name:
            continue
        kind = classify(_strip_unit(str(name)))
        choices: tuple[str, ...] = ()
        if kind == FieldKind.SEVERITY:
            choices = SEVERITY_CHOICES
        elif kind == FieldKind.DISEASE_PRESENCE:
            choices = PRESENCE_CHOICES
        elif kind == FieldKind.RATING:
            choices = RATING_CHOICES
        fields.append(Field(name=str(name), kind=kind, excel_col=col, choices=choices))
    wb.close()
    return fields


# ---- Field grouping for UI layout -------------------------------------------

def pair_disease_fields(fields: list[Field]) -> list[tuple[Field, Field | None]]:
    """Pair each `Disease_X` with its `Disease_X_Severity`.

    Returns rows for a (Disease | Presence | Severity) table.
    """
    by_key = {f.key: f for f in fields}
    pairs: list[tuple[Field, Field | None]] = []
    for f in fields:
        if f.kind != FieldKind.DISEASE_PRESENCE:
            continue
        sev = by_key.get(f.key + "_Severity")
        pairs.append((f, sev))
    return pairs


def pair_nutrient_fields(fields: list[Field]) -> list[tuple[Field, Field | None]]:
    """Pair each nutrient (`N`, `P`, ...) with its `<name>_rate` sibling.

    Returns rows for a (Nutrient | Value | Rate) table.
    """
    by_key = {f.key: f for f in fields}
    pairs: list[tuple[Field, Field | None]] = []
    for f in fields:
        if f.kind != FieldKind.NUMBER:
            continue
        k = f.key
        # Skip suffixed fields; we look for the bare nutrient key and
        # match it against its `_rate` sibling.
        if (
            k.endswith("_rate")
            or k.endswith("_Actual")
            or k.endswith("_Expected")
            or k.startswith("TDR_")
            or k.startswith("Petal_Test_")
        ):
            continue
        rate = by_key.get(k + "_rate")
        if rate is not None:
            pairs.append((f, rate))
    return pairs


def pair_ratio_fields(
    fields: list[Field],
) -> list[tuple[str, Field, Field | None]]:
    """Pair each `<X>_Actual` with `<X>_Expected`.

    Returns rows for a (Ratio | Actual | Expected) table.
    The first element is the bare ratio name (e.g. `N/S`).
    """
    by_key = {f.key: f for f in fields}
    rows: list[tuple[str, Field, Field | None]] = []
    for f in fields:
        if not f.key.endswith("_Actual"):
            continue
        base = f.key[: -len("_Actual")]
        rows.append((base, f, by_key.get(base + "_Expected")))
    return rows


def tdr_fields(fields: list[Field]) -> list[Field]:
    return [f for f in fields if f.key.startswith("TDR_")]


def petal_test_fields(fields: list[Field]) -> list[Field]:
    return [f for f in fields if f.key.startswith("Petal_Test_")]


def read_template_locations(template_path: Path) -> list[tuple[str, float, float]]:
    """Return [(location_id, lat, lon), ...] from col A + col B of data rows."""
    wb = openpyxl.load_workbook(template_path, read_only=True, data_only=True)
    ws = wb.active
    out: list[tuple[str, float, float]] = []
    for row in range(2, ws.max_row + 1):
        loc_id = ws.cell(row=row, column=1).value
        loc_str = ws.cell(row=row, column=2).value
        if not loc_id or not loc_str:
            continue
        lat_str, lon_str = (p.strip() for p in str(loc_str).split(","))
        out.append((str(loc_id), float(lat_str), float(lon_str)))
    wb.close()
    return out
